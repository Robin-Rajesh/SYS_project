"""
export_eval_report.py — Export SQL Eval Results to Excel & PDF
==============================================================
For each prompt in eval_sql_dataset.json, this script:
  1. Runs the prompt through the AI agent to get the generated SQL
  2. Executes both the proposed (expected) SQL and the generated SQL
  3. Captures their outputs (up to 10 rows)
  4. Exports everything to:
       - outputs/eval_report.xlsx
       - outputs/eval_report.pdf

Usage:
  python evals/export_eval_report.py

Dependencies (should already be installed):
  pip install openpyxl reportlab
"""

import os
import sys
import json
import re
import textwrap
from datetime import datetime

import pandas as pd
from sqlalchemy import text

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from agent import run_agent, clear_memory
from tools.sql_tool import get_engine

# ── Output paths ──────────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)
EXCEL_PATH = os.path.join(OUTPUT_DIR, "eval_report.xlsx")
PDF_PATH   = os.path.join(OUTPUT_DIR, "eval_report.pdf")

MAX_ROWS_PREVIEW = 10


# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_sql(response: str) -> str:
    match = re.search(r"```sql\n(.*?)\n```", response, re.DOTALL | re.IGNORECASE)
    return match.group(1).strip() if match else ""


def run_sql(engine, sql: str):
    if not sql:
        return None, "No SQL provided"
    try:
        with engine.connect() as conn:
            df = pd.read_sql(text(sql), conn)
        return df, None
    except Exception as e:
        return None, str(e)


def df_to_str(df, max_rows=MAX_ROWS_PREVIEW) -> str:
    if df is None:
        return "(not executed)"
    if df.empty:
        return "(0 rows returned)"
    preview = df.head(max_rows)
    lines = [" | ".join(str(c) for c in preview.columns)]
    lines.append("-" * len(lines[0]))
    for _, row in preview.iterrows():
        lines.append(" | ".join(str(v) for v in row))
    extra = len(df) - max_rows
    if extra > 0:
        lines.append(f"... and {extra} more rows")
    lines.append(f"(Total rows: {len(df)})")
    return "\n".join(lines)


def results_match(df_expected, df_generated) -> str:
    if df_expected is None or df_generated is None:
        return "ERROR"
    try:
        exp = {tuple(str(v) for v in row)
               for row in df_expected.itertuples(index=False, name=None)}
        gen = {tuple(str(v) for v in row)
               for row in df_generated.itertuples(index=False, name=None)}
        return "PASS" if exp == gen else "FAIL"
    except Exception:
        return "ERROR"


# ── Core collection loop ───────────────────────────────────────────────────────

def collect_results(dataset_path: str) -> list:
    with open(dataset_path, "r", encoding="utf-8") as f:
        dataset = json.load(f)

    engine = get_engine()
    records = []

    print(f"\nRunning {len(dataset)} prompts through the agent ...\n")

    for i, item in enumerate(dataset):
        question     = item["question"]
        expected_sql = item["expected_sql"]

        print(f"  [{i+1}/{len(dataset)}] {question[:80]}...")

        clear_memory()
        response      = run_agent(question)
        generated_sql = extract_sql(response)

        df_exp, err_exp = run_sql(engine, expected_sql)
        df_gen, err_gen = run_sql(engine, generated_sql)

        match_status = results_match(df_exp, df_gen)

        records.append({
            "Q#":                   i + 1,
            "Prompt":               question,
            "Proposed SQL":         expected_sql,
            "Generated SQL":        generated_sql if generated_sql else "(none generated)",
            "Proposed SQL Output":  df_to_str(df_exp) if df_exp is not None else f"ERROR: {err_exp}",
            "Generated SQL Output": df_to_str(df_gen) if df_gen is not None else f"ERROR: {err_gen}",
            "Results Match":        match_status,
            "_df_exp":              df_exp,
            "_df_gen":              df_gen,
        })

    print(f"\nDone. Collected {len(records)} records.\n")
    return records


# ── Excel export ───────────────────────────────────────────────────────────────

def export_excel(records: list, path: str):
    print(f"Writing Excel -> {path}")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    summary_rows = []
    for r in records:
        summary_rows.append({
            "Q#":                   r["Q#"],
            "Prompt":               r["Prompt"],
            "Proposed SQL":         r["Proposed SQL"],
            "Generated SQL":        r["Generated SQL"],
            "Proposed SQL Output":  r["Proposed SQL Output"],
            "Generated SQL Output": r["Generated SQL Output"],
            "Results Match":        r["Results Match"],
        })

    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_widths = {"A": 5, "B": 50, "C": 60, "D": 60, "E": 50, "F": 50, "G": 14}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        pass_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")
        err_fill  = PatternFill("solid", fgColor="FFEB9C")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border    = border
            match_cell = row[6]
            val = str(match_cell.value)
            if "PASS" in val:
                match_cell.fill = pass_fill
            elif "FAIL" in val:
                match_cell.fill = fail_fill
            else:
                match_cell.fill = err_fill

        ws.row_dimensions[1].height = 30

        for r in records:
            q = r["Q#"]
            for label, df in [("Proposed", r["_df_exp"]), ("Generated", r["_df_gen"])]:
                sheet_name = f"Q{q}_{label[:3]}"[:31]
                if df is not None and not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws2 = writer.sheets[sheet_name]
                    for cell in ws2[1]:
                        cell.fill      = PatternFill("solid", fgColor="2E4057")
                        cell.font      = Font(bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center")
                    for col in ws2.columns:
                        max_len = max(
                            len(str(cell.value)) if cell.value else 0
                            for cell in col
                        )
                        ws2.column_dimensions[
                            get_column_letter(col[0].column)
                        ].width = min(max_len + 4, 40)

    print(f"  Excel saved: {path}\n")


# ── PDF export ─────────────────────────────────────────────────────────────────

def export_pdf(records: list, path: str):
    """
    Generates a PDF validation report using ReportLab.
    Uses a stacked (vertical) layout so content of any length fits correctly.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable, KeepTogether,
        )
    except ImportError:
        print("  reportlab not installed. Run: pip install reportlab")
        return

    print(f"Writing PDF  -> {path}")

    PAGE = A4
    W, _H = PAGE

    DOC = SimpleDocTemplate(
        path,
        pagesize=PAGE,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,   bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    usable_width = W - 3*cm

    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#1F3864"), spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        "QHead", parent=styles["Heading2"],
        fontSize=10, textColor=colors.HexColor("#2E4057"),
        spaceBefore=6, spaceAfter=3,
    )
    label_style = ParagraphStyle(
        "SLabel", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#333333"),
        spaceBefore=4, spaceAfter=1,
    )
    mono_style = ParagraphStyle(
        "MonoBlock", parent=styles["Code"],
        fontSize=6.5, leading=8.5,
        backColor=colors.HexColor("#F7F7F7"),
        leftIndent=4, rightIndent=4,
        spaceBefore=1, spaceAfter=3,
    )
    normal_style = ParagraphStyle(
        "NormalTxt", parent=styles["Normal"],
        fontSize=9, leading=12,
    )

    PASS_COLOR = colors.HexColor("#C6EFCE")
    FAIL_COLOR = colors.HexColor("#FFC7CE")
    ERR_COLOR  = colors.HexColor("#FFEB9C")
    HDR_COLOR  = colors.HexColor("#1F3864")

    def safe_para(text: str, max_chars: int = 2000) -> str:
        """Escape XML chars and truncate for safe use inside Paragraph."""
        if len(text) > max_chars:
            text = text[:max_chars] + "\n... [truncated]"
        return (text
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br/>"))

    def section(label: str, content: str) -> list:
        return [
            Paragraph(label, label_style),
            Paragraph(safe_para(content), mono_style),
        ]

    story = []

    # Cover page
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("AI Analytics - SQL Evaluation Report", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Total prompts: <b>{len(records)}</b>",
        normal_style,
    ))
    story.append(Spacer(1, 0.5*cm))

    passed = sum(1 for r in records if r["Results Match"] == "PASS")
    failed = sum(1 for r in records if r["Results Match"] == "FAIL")
    errors = len(records) - passed - failed

    cov_data = [
        ["Total Prompts", "Passed", "Failed", "Errors"],
        [str(len(records)), str(passed), str(failed), str(errors)],
    ]
    cov_tbl = Table(cov_data, colWidths=[usable_width / 4] * 4)
    cov_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  HDR_COLOR),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 11),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND",    (1, 1), (1, 1),   PASS_COLOR),
        ("BACKGROUND",    (2, 1), (2, 1),   FAIL_COLOR),
        ("BACKGROUND",    (3, 1), (3, 1),   ERR_COLOR),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
    ]))
    story.append(cov_tbl)
    story.append(PageBreak())

    # Per-question pages — pure stacked layout, no side-by-side tables
    for r in records:
        match = r["Results Match"]
        if match == "PASS":
            badge_color = PASS_COLOR
            badge_text  = "PASS"
        elif match == "FAIL":
            badge_color = FAIL_COLOR
            badge_text  = "FAIL"
        else:
            badge_color = ERR_COLOR
            badge_text  = "ERROR"

        badge_tbl = Table([[badge_text]], colWidths=[3*cm])
        badge_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), badge_color),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME",      (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ]))

        # Header block kept together
        story.append(KeepTogether([
            Paragraph(f"Q{r['Q#']}: {safe_para(r['Prompt'])}", h2_style),
            HRFlowable(width="100%", thickness=0.8, color=colors.HexColor("#BBBBBB")),
            Spacer(1, 0.15*cm),
            badge_tbl,
            Spacer(1, 0.2*cm),
        ]))

        # Stacked sections — ReportLab paginates these freely
        story.extend(section("Proposed SQL (Expected):", r["Proposed SQL"]))
        story.extend(section("Proposed SQL Output:", r["Proposed SQL Output"]))

        story.append(HRFlowable(width="100%", thickness=0.4,
                                color=colors.HexColor("#DDDDDD")))
        story.append(Spacer(1, 0.1*cm))

        story.extend(section("Generated SQL (AI Output):", r["Generated SQL"]))
        story.extend(section("Generated SQL Output:", r["Generated SQL Output"]))

        story.append(PageBreak())

    DOC.build(story)
    print(f"  PDF saved: {path}\n")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    dataset_file = os.path.join(os.path.dirname(__file__), "eval_sql_dataset.json")

    print("=" * 65)
    print("  SQL EVALUATION REPORT EXPORTER")
    print("=" * 65)

    records = collect_results(dataset_file)

    export_excel(records, EXCEL_PATH)
    export_pdf(records, PDF_PATH)

    print("=" * 65)
    print(f"  [Excel] {EXCEL_PATH}")
    print(f"  [PDF]   {PDF_PATH}")
    print("=" * 65)


if __name__ == "__main__":
    main()
