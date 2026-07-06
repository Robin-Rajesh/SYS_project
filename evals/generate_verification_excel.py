"""
generate_verification_excel.py — Output Verification Report
============================================================
Reads eval_report.xlsx and produces a clean verification Excel
focused purely on comparing AI-generated output vs actual backend data.

Columns per sheet:
  Summary  : Q#, Prompt, Match Status, Row Count (Expected), Row Count (Generated)
  Q1, Q2.. : Side-by-side actual backend rows vs AI generated rows

Usage:
  python evals/generate_verification_excel.py
"""

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "..", "outputs")
EXCEL_IN    = os.path.join(OUTPUT_DIR, "eval_report.xlsx")
EXCEL_OUT   = os.path.join(OUTPUT_DIR, "verification_report.xlsx")


# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E4057"
PASS_GREEN  = "C6EFCE"
FAIL_RED    = "FFC7CE"
ERR_YELLOW  = "FFEB9C"
HEADER_FG   = "FFFFFF"
ALT_ROW     = "EBF0FA"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_cell(ws, row, col, value, bg=DARK_BLUE, fg=HEADER_FG, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
    return c


def data_cell(ws, row, col, value, bg=None, bold=False, wrap=True):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, size=9)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border    = thin_border()
    return c


def parse_output_to_df(text: str) -> pd.DataFrame:
    """
    Parse the pipe-delimited text output stored in the Excel back into a DataFrame.
    Returns empty DataFrame on failure.
    """
    text = str(text).strip()
    lines = [l for l in text.split("\n") if l.strip()]
    # Filter out footnotes and dividers
    data_lines = [l for l in lines
                  if not l.strip().startswith("-")
                  and not l.strip().startswith("(")
                  and not l.strip().startswith("...")]

    if len(data_lines) < 2:
        return pd.DataFrame()

    try:
        headers = [h.strip() for h in data_lines[0].split("|")]
        rows = []
        for line in data_lines[1:]:
            cells = [c.strip() for c in line.split("|")]
            if len(cells) == len(headers):
                rows.append(cells)
        return pd.DataFrame(rows, columns=headers)
    except Exception:
        return pd.DataFrame()


def build_verification_excel(summary_df: pd.DataFrame, path: str):
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Verification Summary"

    headers = ["Q#", "Prompt", "Match Status",
               "Rows (Expected)", "Rows (Generated)", "Notes"]
    col_widths = [5, 65, 16, 16, 16, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hdr_cell(ws_sum, 1, col, h)
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    ws_sum.row_dimensions[1].height = 28

    for i, row in summary_df.iterrows():
        r = i + 2
        match = str(row.get("Results Match", "")).strip().upper()

        if match == "PASS":
            badge_bg = PASS_GREEN
        elif match == "FAIL":
            badge_bg = FAIL_RED
        else:
            badge_bg  = ERR_YELLOW
            match     = "ERROR"

        bg = ALT_ROW if r % 2 == 0 else None

        # Count rows from parsed output text
        exp_df = parse_output_to_df(str(row.get("Proposed SQL Output",  "")))
        gen_df = parse_output_to_df(str(row.get("Generated SQL Output", "")))
        exp_rows = len(exp_df) if not exp_df.empty else "N/A"
        gen_rows = len(gen_df) if not gen_df.empty else "N/A"

        note = ""
        if match == "FAIL":
            note = "Outputs differ — check per-question sheet"
        elif match == "PASS":
            note = "Outputs match"

        data_cell(ws_sum, r, 1, row.get("Q#", ""), bg=bg, bold=True)
        data_cell(ws_sum, r, 2, row.get("Prompt", ""), bg=bg)
        c = ws_sum.cell(row=r, column=3, value=match)
        c.fill      = PatternFill("solid", fgColor=badge_bg)
        c.font      = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        data_cell(ws_sum, r, 4, exp_rows, bg=bg)
        data_cell(ws_sum, r, 5, gen_rows, bg=bg)
        data_cell(ws_sum, r, 6, note, bg=bg)

    ws_sum.freeze_panes = "A2"

    # ── Per-question sheets: side-by-side comparison ──────────────────────────
    for i, row in summary_df.iterrows():
        q_num = row.get("Q#", i + 1)
        match = str(row.get("Results Match", "")).strip().upper()

        exp_df = parse_output_to_df(str(row.get("Proposed SQL Output",  "")))
        gen_df = parse_output_to_df(str(row.get("Generated SQL Output", "")))

        if exp_df.empty and gen_df.empty:
            continue

        sheet_name = f"Q{q_num}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Question prompt as title
        prompt_cell = ws.cell(row=1, column=1,
                              value=f"Q{q_num}: {row.get('Prompt', '')}")
        prompt_cell.font      = Font(bold=True, size=11, color=MID_BLUE)
        prompt_cell.alignment = Alignment(wrap_text=True)

        # Merge across expected + generated columns
        n_exp_cols = len(exp_df.columns) if not exp_df.empty else 3
        n_gen_cols = len(gen_df.columns) if not gen_df.empty else 3
        total_cols = n_exp_cols + 1 + n_gen_cols  # +1 gap column

        if total_cols > 1:
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=total_cols)

        ws.row_dimensions[1].height = 40

        # Match status banner
        status_text = f"Result: {match}"
        status_bg   = PASS_GREEN if match == "PASS" else (FAIL_RED if match == "FAIL" else ERR_YELLOW)
        banner = ws.cell(row=2, column=1, value=status_text)
        banner.fill      = PatternFill("solid", fgColor=status_bg)
        banner.font      = Font(bold=True, size=10)
        banner.alignment = Alignment(horizontal="center")
        if total_cols > 1:
            ws.merge_cells(start_row=2, start_column=1,
                           end_row=2, end_column=total_cols)

        # Section headers
        hdr_cell(ws, 3, 1, "EXPECTED OUTPUT (Actual Backend Data)",
                 bg=DARK_BLUE, size=9)
        if n_exp_cols > 1:
            ws.merge_cells(start_row=3, start_column=1,
                           end_row=3, end_column=n_exp_cols)

        gap_col = n_exp_cols + 1
        ws.cell(row=3, column=gap_col, value="")  # gap

        hdr_cell(ws, 3, gap_col + 1, "AI GENERATED OUTPUT",
                 bg=MID_BLUE, size=9)
        if n_gen_cols > 1:
            ws.merge_cells(start_row=3, start_column=gap_col + 1,
                           end_row=3, end_column=gap_col + n_gen_cols)

        # Column headers — expected
        if not exp_df.empty:
            for c_idx, col_name in enumerate(exp_df.columns, start=1):
                hdr_cell(ws, 4, c_idx, col_name, bg="344E6E", size=8)
                ws.column_dimensions[get_column_letter(c_idx)].width = 18

        # Column headers — generated
        if not gen_df.empty:
            for c_idx, col_name in enumerate(gen_df.columns,
                                             start=gap_col + 1):
                hdr_cell(ws, 4, c_idx, col_name, bg="3D5A80", size=8)
                ws.column_dimensions[get_column_letter(c_idx)].width = 18

        ws.column_dimensions[get_column_letter(gap_col)].width = 3

        # Data rows — sorted for easy comparison
        exp_sorted = (exp_df.sort_values(by=list(exp_df.columns))
                      .reset_index(drop=True) if not exp_df.empty else pd.DataFrame())
        gen_sorted = (gen_df.sort_values(by=list(gen_df.columns))
                      .reset_index(drop=True) if not gen_df.empty else pd.DataFrame())

        max_rows = max(len(exp_sorted), len(gen_sorted))
        for r_idx in range(max_rows):
            excel_row = r_idx + 5
            row_bg = ALT_ROW if r_idx % 2 == 0 else None

            # Expected data
            if r_idx < len(exp_sorted):
                for c_idx, val in enumerate(exp_sorted.iloc[r_idx], start=1):
                    data_cell(ws, excel_row, c_idx, val, bg=row_bg)
            else:
                for c_idx in range(1, n_exp_cols + 1):
                    data_cell(ws, excel_row, c_idx, "", bg=row_bg)

            # Gap column
            ws.cell(row=excel_row, column=gap_col, value="")

            # Generated data
            if r_idx < len(gen_sorted):
                for c_idx, val in enumerate(gen_sorted.iloc[r_idx],
                                            start=gap_col + 1):
                    # Highlight mismatches if same column count
                    cell_bg = row_bg
                    if (not exp_sorted.empty and r_idx < len(exp_sorted)
                            and c_idx - gap_col - 1 < len(exp_sorted.columns)
                            and c_idx - gap_col - 1 < len(gen_sorted.columns)):
                        exp_val = str(exp_sorted.iloc[r_idx, c_idx - gap_col - 1])
                        gen_val = str(val)
                        if exp_val != gen_val:
                            cell_bg = FAIL_RED
                    data_cell(ws, excel_row, c_idx, val, bg=cell_bg)
            else:
                for c_idx in range(gap_col + 1, gap_col + n_gen_cols + 1):
                    data_cell(ws, excel_row, c_idx, "", bg=FAIL_RED)

        ws.freeze_panes = "A5"

    wb.save(path)
    print(f"Verification Excel saved: {path}")


def main():
    if not os.path.exists(EXCEL_IN):
        print(f"ERROR: {EXCEL_IN} not found.")
        print("Run export_eval_report.py first.")
        sys.exit(1)

    print(f"Reading: {EXCEL_IN}")
    df = pd.read_excel(EXCEL_IN, sheet_name="Summary")
    print(f"Loaded {len(df)} rows.")
    print(f"Writing: {EXCEL_OUT}")
    build_verification_excel(df, EXCEL_OUT)


if __name__ == "__main__":
    main()
